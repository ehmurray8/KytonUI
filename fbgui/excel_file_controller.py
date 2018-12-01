import queue
import traceback
from tkinter import messagebox
from typing import Tuple, Callable

import pandas as pd
from StyleFrame import Styler, utils, StyleFrame
from openpyxl import drawing
from openpyxl.chart import ScatterChart, Reference, Series, marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.colors import ColorChoice, RGBPercent
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties
from openpyxl.utils import get_column_letter

from fbgui.baking_curve_fit import curve_fit_baking, add_baking_trend_line, create_sensitivity_line
from fbgui.calibration_excel_container import CalibrationExcelContainer
from fbgui.data_container import DataCollection
from fbgui.database_controller import DatabaseController
from fbgui.excel_graph_helpers import *
from fbgui.helpers import get_file_name
from fbgui.messages import *


class ExcelFileController:
    def __init__(self, excel_file_path: str, fbg_names: List[str], main_queue: queue.Queue, function_type: str,
                 bake_sensitivity: float=None):
        self.excel_file_path = excel_file_path
        self.excel_file_name = get_file_name(excel_file_path)
        self.fbg_names = fbg_names
        self.main_queue = main_queue
        self.bake_sensitivity = bake_sensitivity
        self.database_controller = DatabaseController(excel_file_path, fbg_names, main_queue, function_type)
        self.is_calibration = function_type == CAL

    def create_excel(self):
        try:
            if self.is_calibration:
                self.create_calibration_excel()
            else:
                self.create_baking_excel()
        except (RuntimeError, IndexError):
            self.main_queue.put(Message(MessageType.WARNING, "Excel File Creation Error",
                                        "No data has been recorded yet, or the database has been corrupted."))
            traceback.print_exc()
        except PermissionError:
            message = "Please close {}, before attempting to create a new copy of it.".format(self.excel_file_name)
            self.main_queue.put(Message(MessageType.WARNING, "Excel File Creation Error", message))
            messagebox.showwarning("Excel file is already opened", message)

    def create_baking_excel(self):
        data_frame = self.database_controller.to_data_frame()
        data_collection = DataCollection()
        data_collection.create(self.is_calibration, data_frame, self.fbg_names, self.main_queue)
        column_ordering = [DATE_TIME_HEADER, DELTA_TIME_HEADER, TEMPERATURE_HEADER]

        add_times(data_frame, data_collection)
        add_wavelength_power_columns(column_ordering, data_frame)

        add_baking_duplicate_columns(data_frame, data_collection)
        column_ordering.extend([DELTA_TIME_HEADER1, DELTA_TEMPERATURE_HEADER1])
        self.add_delta_wavelengths(data_frame, data_collection, column_ordering)
        column_ordering.extend([DELTA_TIME_HEADER2, DELTA_TEMPERATURE_HEADER2])
        self.add_delta_powers(data_frame, data_collection, column_ordering)

        column_ordering.extend([DELTA_TEMPERATURE_HEADER3, MEAN_DELTA_WAVELENGTH_HEADER, MEAN_DELTA_POWER_HEADER])
        data_frame[MEAN_DELTA_WAVELENGTH_HEADER] = data_collection.mean_delta_wavelengths_pm
        data_frame[MEAN_DELTA_POWER_HEADER] = data_collection.mean_delta_powers
        data_frame = data_frame[column_ordering]

        baking_coefficients = self._create_baking_coefficients(data_frame)
        trend_line_indexes = self._create_trend_lines(data_frame, baking_coefficients)
        sensitivity_indexes = self._create_sensitivity_lines(data_frame, trend_line_indexes)

        style_frame = self.create_style_frame(data_frame)
        style_frame.apply_column_style(cols_to_style=MEAN_DELTA_WAVELENGTH_HEADER,
                                       styler_obj=Styler(font_color=utils.colors.red))
        style_frame.apply_column_style(cols_to_style=[DATE_TIME_HEADER],
                                       styler_obj=Styler(number_format=utils.number_formats.date_time_with_seconds))
        delta_temperature_headers = [col for col in data_frame.columns.values if DELTA_TEMPERATURE_HEADER in col]
        style_frame.apply_column_style(cols_to_style=delta_temperature_headers,
                                       styler_obj=Styler(font_color=utils.colors.red))
        parameters = BakingGraphParameters(len(data_frame.index), trend_line_indexes, sensitivity_indexes)
        self.show_excel([style_frame], ["Baking Data"], parameters, self._graph_bake_results,
                        baking_coefficients=baking_coefficients)

    def _create_baking_coefficients(self, data_frame: pd.DataFrame) -> List[List[float]]:
        baking_coefficients = []
        for i in range(len(self.fbg_names)):
            baking_coefficients.append(curve_fit_baking(data_frame, self._get_baking_wavelength_column(i)-1))
        return baking_coefficients

    def _create_trend_lines(self, data_frame: pd.DataFrame, coefficients: List[List[float]]) -> List[int]:
        trend_line_indexes = []
        for i, fbg_name in enumerate(self.fbg_names):
            trend_line_indexes.append(add_baking_trend_line(data_frame, list(reversed(coefficients[i])), fbg_name))
        return trend_line_indexes

    def _create_sensitivity_lines(self, data_frame: pd.DataFrame, trend_line_indexes: List[int]) -> List[int]:
        sensitivity_indexes = []
        if self.bake_sensitivity is not None:
            for trend_index, fbg_name in zip(trend_line_indexes, self.fbg_names):
                sensitivity_index = create_sensitivity_line(data_frame, self.bake_sensitivity, trend_index, fbg_name)
                sensitivity_indexes.append(sensitivity_index)
        return sensitivity_indexes

    def create_calibration_excel(self):
        data_frame = self.database_controller.to_data_frame()
        data_collection = DataCollection()
        data_collection.create(self.is_calibration, data_frame, self.fbg_names, self.main_queue)

        real_point_data_frame = data_frame[data_frame[REAL_POINT_HEADER] == "True"]
        cycles = list(set(real_point_data_frame[CYCLE_HEADER]))
        cycles.sort()
        del real_point_data_frame[REAL_POINT_HEADER]
        del real_point_data_frame[DATE_TIME_HEADER]

        full_column_ordering = [DATE_TIME_HEADER, DELTA_TIME_HEADER, TEMPERATURE_HEADER]

        add_times(data_frame, data_collection)
        add_wavelength_power_columns(full_column_ordering, data_frame)
        full_column_ordering.extend([CYCLE_HEADER, REAL_POINT_HEADER])
        data_frame = data_frame[full_column_ordering]

        container = CalibrationExcelContainer(real_point_data_frame, cycles)
        calibration_data_frame = container.get_data_frame()
        calibration_style_frame = self.create_style_frame(calibration_data_frame)
        full_style_frame = self.create_style_frame(data_frame)
        delta_temperature_headers = [col for col in data_frame.columns.values if DELTA_TEMPERATURE_HEADER in col]
        full_style_frame.apply_column_style(cols_to_style=delta_temperature_headers,
                                            styler_obj=Styler(font_color=utils.colors.red))
        full_style_frame.apply_column_style(cols_to_style=[DATE_TIME_HEADER],
                                            styler_obj=
                                            Styler(number_format=utils.number_formats.date_time_with_seconds))
        first_column = "Temperature (K) Cycle {}".format(container.cycles[0])
        first_temp = calibration_data_frame[first_column].values[0]
        last_temp = calibration_data_frame[first_column].values[-1] + 5
        parameters = CalibrationGraphParameters(len(calibration_data_frame.index), [first_temp, last_temp],
                                                container.cycles, container.mean_wavelength_indexes,
                                                container.deviation_wavelength_indexes, container.mean_power_indexes,
                                                container.deviation_power_indexes,
                                                container.sensitivity_wavelength_indexes,
                                                container.sensitivity_power_indexes,
                                                container.master_temperature_column, container.mean_temperature_column)
        coefficients = [container.wavelength_mean_coefficients, container.power_mean_coefficients]
        self.show_excel([calibration_style_frame, full_style_frame], ["Cal", "Full Cal"],
                        parameters, self._graph_calibration_results, coefficients)

    def write_curve_fit_coefficients(self, coefficients_list: List[Tuple[List[List[float]], List[List[float]]]],
                                     worksheet: Worksheet):
        if not len(coefficients_list[0]) or not len(coefficients_list[1]):
            return

        headers = ["Wavelength Deviation: Fit parameters from mean calibration curve",
                   "Power Deviation: Fit parameters from mean calibration curve"]
        row = 1
        for header, (coefficients, derivative_coefficients) in zip(headers, coefficients_list):
            worksheet.append([header])
            worksheet.merge_cells("A{}:H{}".format(row, row))
            row += 1
            row = self.write_coefficients(worksheet, row, coefficients, derivative_coefficients, row == 2)
            for _ in range(0, 3):
                worksheet.append([])
                row += 1

    def write_coefficients(self, worksheet: Worksheet, row: int, coefficients: List[List[float]],
                           derivative_coefficients: List[List[float]]=None, write_sensitivity: bool = False) -> int:
        sensitivity_header = "Sensitivity (pm/K)"
        drift_rate_header = "Drift Rate (mK/hr)"
        for i, name in enumerate(self.fbg_names):
            worksheet.merge_cells("A{}:E{}".format(row, row))
            values = [name]
            if derivative_coefficients is None and row == 1:
                values.extend(["", "", "", "", sensitivity_header, "", drift_rate_header])
                worksheet.column_dimensions["F"].width = len(sensitivity_header)
                worksheet.column_dimensions["H"].width = len(drift_rate_header)
            worksheet.append(values)
            row += 1

            coefficient_specifiers = ["x^{}".format(num) for num in reversed(range(0, len(coefficients[i])))]
            offset = (5 - len(list(filter(lambda x: x != 0, coefficients[i]))))
            coefficient_specifiers.extend([""] * offset)
            if derivative_coefficients is not None:
                coefficient_specifiers\
                    .extend(["x^{}".format(num) for num in reversed(range(0, len(derivative_coefficients[i])))])
                if write_sensitivity:
                    coefficient_specifiers.extend([""] * offset)
                    temperature_header = "Temperature [C]"
                    coefficient_specifiers.extend([temperature_header, sensitivity_header])
                    column = get_column_letter(coefficient_specifiers.index(temperature_header) + 1)
                    worksheet.column_dimensions[column].width = len(temperature_header) + 1
                    column = get_column_letter(coefficient_specifiers.index(sensitivity_header) + 1)
                    worksheet.column_dimensions[column].width = len(sensitivity_header) + 1
            worksheet.append(coefficient_specifiers)
            row += 1

            row_values = []
            row_values.extend(coefficients[i])
            row_values.extend([""] * (5 - len(list(filter(lambda x: x != 0, coefficients[i])))))
            if derivative_coefficients is not None:
                row_values.extend(derivative_coefficients[i])
                if write_sensitivity:
                    row_values.extend([""] * offset)
                    row_values.append(100)
                    start = len(coefficients[i]) + offset + 1
                    temperature_index = start + len(derivative_coefficients[i]) + offset
                    temperature_column = get_column_letter(temperature_index)
                    columns = [get_column_letter(i) for i in range(start, start + len(derivative_coefficients[i]))]
                    temperature_identifier = "{}{}".format(temperature_column, row)
                    equation_parts = []
                    for k, column in enumerate(columns):
                        exponent = len(columns) - 1 - k
                        equation_parts.append("(({} + 273.15)^{} * {}{})"
                                              .format(temperature_identifier, exponent, column, row))
                    equation = "=({}) * 1000".format(" + ".join(equation_parts))
                    row_values.append(equation)
            else:
                sensitivity = 1
                if self.bake_sensitivity is not None:
                    sensitivity = self.bake_sensitivity
                row_values.extend([sensitivity, "",
                                   '=IF(ISNUMBER(A{0}), IF(ISNUMBER(F{0}),  A{0} * 1000 / F{0}, ""), "")' .format(row)])
            worksheet.append(row_values)
            row += 1
        return row

    def add_delta_wavelengths(self, data_frame: pd.DataFrame, data_collection: DataCollection,
                              column_ordering: List[str]):
        delta_wavelengths = data_collection.delta_wavelengths_pm
        for fbg_name, delta_wave in zip(self.fbg_names, delta_wavelengths):
            delta_wavelength_header = "{} {}{} (pm.)".format(fbg_name, u"\u0394", u"\u03BB")
            data_frame[delta_wavelength_header] = delta_wave
            column_ordering.append(delta_wavelength_header)

    def add_delta_powers(self, data_frame: pd.DataFrame, data_collection: DataCollection, column_ordering: List[str]):
        delta_powers = data_collection.delta_powers
        for fbg_name, delta_power in zip(self.fbg_names, delta_powers):
            delta_power_header = "{} {}{} (dBm.)".format(fbg_name, u"\u0394", "P")
            data_frame[delta_power_header] = delta_power
            column_ordering.append(delta_power_header)

    def create_style_frame(self, data_frame: pd.DataFrame):
        defaults = {'font_size': 14}
        style_frame = StyleFrame(data_frame, styler_obj=Styler(**defaults, shrink_to_fit=False, wrap_text=False))

        header_style = Styler(bold=True, font_size=18)
        style_frame.set_column_width(columns=style_frame.columns, width=35)
        style_frame.apply_headers_style(styler_obj=header_style)

        for fbg_name, hex_color in zip(self.fbg_names, HEX_COLORS):
            style_frame.apply_column_style(cols_to_style=[col for col in data_frame.columns.values if fbg_name in col],
                                           styler_obj=Styler(bg_color=hex_color))
        temperature_headers = [col for col in data_frame.columns.values if "Temperature" in col]
        style_frame.apply_column_style(cols_to_style=temperature_headers,
                                       styler_obj=Styler(font_color=utils.colors.red))
        return style_frame

    def show_excel(self, style_frames: List[StyleFrame], sheet_names: List[str], parameters: GraphParameters,
                   graph_results: Callable[[GraphParameters], None],
                   coefficients: List[Tuple[List[List[float]], List[List[float]]]]=None,
                   baking_coefficients: List[List[float]]=None):
        excel_writer = StyleFrame.ExcelWriter(self.excel_file_path)
        for style_frame, sheet_name in zip(style_frames, sheet_names):
            style_frame.to_excel(excel_writer=excel_writer, row_to_add_filters=0, sheet_name=sheet_name)

        if coefficients is not None:
            worksheet = excel_writer.book.create_sheet("Curve Fit")
            self.write_curve_fit_coefficients(coefficients, worksheet)
        if baking_coefficients is not None:
            worksheet = excel_writer.book.create_sheet("Curve Fit")
            self.write_coefficients(worksheet, 1, baking_coefficients)

        parameters.chart_sheet = excel_writer.book.create_sheet(title="Chart")
        parameters.data_sheet = excel_writer.sheets[sheet_names[0]]
        graph_results(parameters)

        excel_writer.book.create_sheet("Notes")

        excel_writer.save()
        excel_writer.close()
        os.startfile('"{}"'.format(self.excel_file_path.replace("\\", "\\\\")))

    def _graph_calibration_results(self, calibration_parameters: CalibrationGraphParameters):
        self.graph_wavelength_deviation(calibration_parameters)
        self.graph_wavelength_means(calibration_parameters)
        self.graph_power_deviation(calibration_parameters)
        self.graph_power_means(calibration_parameters)
        self.graph_wavelength_sensitivity(calibration_parameters)
        self.graph_power_sensitivity(calibration_parameters)

    def _graph_bake_results(self, parameters: BakingGraphParameters):
        self.graph_bake(parameters)

    def graph_bake(self, parameters: BakingGraphParameters):
        last_row = parameters.num_rows + 1
        start_row = 2
        start_column = 2

        x_axis_title = "{} Time from start (hr)".format(u"\u0394")
        y_axis_title = "{} Wavelength (pm)".format(u"\u0394")
        y_axis_title_sensitivity = "Drift (mK)"
        x_values = Reference(parameters.data_sheet, min_col=start_column, min_row=start_row, max_row=last_row)
        for i, fbg_name in enumerate(self.fbg_names):
            chart_title = "{} {} Wavelength (pm) vs. {} Time from start ; {}"\
                .format(fbg_name, u"\u0394", u"\u0394", self.excel_file_name)
            chart = ScatterChart()
            y_values = Reference(parameters.data_sheet, min_col=self._get_baking_wavelength_column(i),
                                 min_row=start_row, max_row=last_row)
            series = self.create_series_bake(x_values, y_values, i)
            chart.series.append(series)

            trend_values = Reference(parameters.data_sheet, min_col=parameters.trend_line_indexes[i] + 1,
                                     min_row=start_row, max_row=last_row)
            series = Series(trend_values, x_values, title="Trend (pm)")
            chart.series.append(series)
            format_chart(chart, x_axis_title, y_axis_title, chart_title)
            parameters.chart_sheet.add_chart(chart, "B{}".format(30*i + 2))

            if self.bake_sensitivity is not None:
                sensitivity_chart_title = "{} drift (mK) vs. {} Time from start; {}"\
                    .format(fbg_name, u"\u0394", self.excel_file_name)
                sensitivity_chart = ScatterChart()
                sensitivity_values = Reference(parameters.data_sheet, min_col=parameters.sensitivity_indexes[i] + 1,
                                               min_row=start_row, max_row=last_row)
                series = Series(sensitivity_values, x_values, title="Drift (mK)")
                sensitivity_chart.series.append(series)
                format_chart(sensitivity_chart, x_axis_title, y_axis_title_sensitivity, sensitivity_chart_title)
                parameters.chart_sheet.add_chart(sensitivity_chart, "V{}".format(30 * i + 2))

    def _get_baking_wavelength_column(self, fbg_index: int):
        return 2 * len(self.fbg_names) + 5 + fbg_index + 1

    def graph_wavelength_deviation(self, parameters: CalibrationGraphParameters):
        self._graph_calibration(parameters.deviation_wavelength_indexes,
                                parameters, CalibrationGraphSubType.WAVELENGTH_DEVIATION, _add_cycle_series,
                                parameters.mean_temperature_column)

    def graph_wavelength_means(self, parameters: CalibrationGraphParameters):
        self._graph_calibration(parameters.mean_wavelength_indexes,
                                parameters, CalibrationGraphSubType.WAVELENGTH_MEAN, _add_mean_series,
                                parameters.mean_temperature_column)

    def graph_power_deviation(self, parameters: CalibrationGraphParameters):
        self._graph_calibration(parameters.deviation_power_indexes, parameters, CalibrationGraphSubType.POWER_DEVIATION,
                                _add_cycle_series, parameters.mean_temperature_column)

    def graph_power_means(self, parameters: CalibrationGraphParameters):
        self._graph_calibration(parameters.mean_power_indexes, parameters, CalibrationGraphSubType.POWER_MEAN,
                                _add_mean_series, parameters.mean_temperature_column)

    def graph_wavelength_sensitivity(self, parameters: CalibrationGraphParameters):
        self._graph_calibration(parameters.sensitivity_wavelength_indexes, parameters,
                                CalibrationGraphSubType.WAVELENGTH_SENSITIVITY, _add_mean_series,
                                parameters.master_temperature_column)

    def graph_power_sensitivity(self, parameters: CalibrationGraphParameters):
        self._graph_calibration(parameters.sensitivity_power_indexes, parameters,
                                CalibrationGraphSubType.POWER_SENSITIVITY, _add_mean_series,
                                parameters.master_temperature_column)

    def _graph_calibration(self, indexes: List[int], parameters: CalibrationGraphParameters,
                           sub_type: CalibrationGraphSubType,
                           add_series: Callable[[ScatterChart, int, int, SeriesParameters], int],
                           static_temperature_column: int):
        series_parameters = SeriesParameters(parameters, indexes, sub_type)
        temperature_column = static_temperature_column
        index = 0
        for i, fbg_name in enumerate(self.fbg_names):
            chart = create_chart(parameters.temperatures)
            index = add_series(chart, index, temperature_column, series_parameters)
            self._create_chart(i, parameters.cycles, chart, parameters.chart_sheet, sub_type)

    def _create_chart(self, index: int, cycles: List[int], chart: ScatterChart, chart_sheet: Worksheet,
                      sub_type: CalibrationGraphSubType):
        chart_title = "{} {} calibration; Run {}; {} cycles" \
            .format(self.fbg_names[index], sub_type.y_axis_title(), self.excel_file_name, len(cycles))
        x_axis_title = "Temperature (K)"
        y_axis_title = sub_type.y_axis_title()
        format_chart(chart, x_axis_title, y_axis_title, chart_title)
        excel_coordinate = sub_type.get_column_letter() \
            + str(sub_type.get_start_row(len(self.fbg_names)) + (index * 30))
        chart_sheet.add_chart(chart, excel_coordinate)

    def create_series_bake(self, x_values: Reference, y_values: Reference, index: int) -> Series:
        series = Series(values=y_values, xvalues=x_values, title=self.fbg_names[index])
        rgb_percent = RGBPercent(*hex_to_rgb(index))
        series.marker = marker.Marker(symbol=MARKERS[index % len(MARKERS)],
                                      spPr=GraphicalProperties(solidFill=ColorChoice(rgb_percent)))
        series.graphicalProperties.line.noFill = True
        return series


def add_times(data_frame: pd.DataFrame, data_collection: DataCollection):
    data_frame[DATE_TIME_HEADER] = data_frame[DATE_TIME_HEADER] \
        .apply(lambda x: x.tz_localize("UTC").tz_convert("US/Eastern").strftime("%m/%d/%y %I:%M %p"))
    data_frame[DELTA_TIME_HEADER] = data_collection.times


def add_wavelength_power_columns(column_ordering: List[str], data_frame: pd.DataFrame):
    wavelength_headers, power_headers = get_wavelength_power_headers(data_frame)
    column_ordering.extend(wavelength_headers)
    column_ordering.extend(power_headers)


def add_baking_duplicate_columns(data_frame: pd.DataFrame, data_collection: DataCollection):
    data_frame[DELTA_TIME_HEADER1] = data_frame[DELTA_TIME_HEADER]
    data_frame[DELTA_TIME_HEADER2] = data_frame[DELTA_TIME_HEADER]
    data_frame[DELTA_TEMPERATURE_HEADER] = data_collection.delta_temps
    data_frame[DELTA_TEMPERATURE_HEADER1] = data_collection.delta_temps
    data_frame[DELTA_TEMPERATURE_HEADER2] = data_collection.delta_temps
    data_frame[DELTA_TEMPERATURE_HEADER3] = data_collection.delta_temps


def get_wavelength_power_headers(data_frame: pd.DataFrame) -> Tuple[List[str], List[str]]:
    headers = data_frame.columns.values.tolist()
    wavelength_headers = [head for head in headers if "Wave" in head]
    power_headers = [head for head in headers if "Pow" in head]
    return wavelength_headers, power_headers


def hex_to_rgb(hex_index: int) -> List[float]:
    hex_string = HEX_COLORS[hex_index][1:]
    rgb_list = []  # type: List[float]
    for i in range(0, len(hex_string), 2):
        hex_int = int(hex_string[i:i + 2], 16) / 255 * 100
        rgb_list.append(hex_int)
    return rgb_list


def format_chart(chart: ScatterChart, x_axis_title: str, y_axis_title: str, title: str):
    chart.height = 15
    chart.width = 30
    chart.x_axis.tickLblPos = "low"

    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title

    font = drawing.text.Font(typeface='Arial')
    cp_axis = CharacterProperties(latin=font, sz=1600, b=True)
    cp_axis_title = CharacterProperties(latin=font, sz=1600)
    cp_title = CharacterProperties(latin=font, sz=1200)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    chart.y_axis.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis),
                                                    endParaRPr=cp_axis_title)])

    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis), endParaRPr=cp_axis)])
    chart.x_axis.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_axis),
                                                    endParaRPr=cp_axis_title)])
    chart.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_title), endParaRPr=cp_title)])


def create_chart(temperatures: List[float]) -> ScatterChart:
    chart = ScatterChart()
    chart.scatterStyle = "smoothMarker"
    chart.x_axis.scaling.min = temperatures[0]
    chart.x_axis.scaling.max = temperatures[-1]
    return chart


def _add_mean_series(chart: ScatterChart, index: int, temperature_column: int,
                     series_parameters: SeriesParameters) -> int:
    return _add_series(chart, index, temperature_column, series_parameters)


def _add_cycle_series(chart: ScatterChart, index: int,
                      temperature_column: int, series_parameters: SeriesParameters) -> int:
    for cycle in series_parameters.cycles:
        index = _add_series(chart, index, temperature_column, series_parameters, cycle)
    return index


def _add_series(chart: ScatterChart, index: int, temperature_column: int,
                series_parameters: SeriesParameters, cycle: int = None):
    x_values = Reference(series_parameters.data_sheet, min_col=temperature_column, min_row=CALIBRATION_START_ROW,
                         max_row=series_parameters.last_row)
    y_values = Reference(series_parameters.data_sheet, min_col=series_parameters.indexes[index] + 1,
                         min_row=CALIBRATION_START_ROW, max_row=series_parameters.last_row)
    series_title = series_parameters.sub_type.get_series_title(cycle)
    series = Series(xvalues=x_values, values=y_values, title=series_title)
    series.marker = marker.Marker(MARKERS[index % len(MARKERS)])
    series.marker.size = 12
    chart.series.append(series)
    return index + 1
